import openpyxl
from openpyxl.styles import colors,Font

wb = openpyxl.Workbook()
ws = wb.active

ws.title = "修改字体样式"

#修改字体样式
ws["C3"].font = Font(name="微软雅黑",size=24,bold=True,italic=True)
ws["C3"] = '微软雅黑 24 加粗 斜体'

ws["A5"].font = Font(name="华文新魏",size=5,bold=False,color=colors.BLUE,italic=False)
ws["A5"] = "华文新魏 5 蓝色"

#为单元格设置公式
ws = wb.create_sheet("设置单元格公式")
ws["B1"].value= 100
ws["B2"].value = 200
ws["B3"].value = "=SUM(B1:B2)"#引号里面应该先写等于号不然的话字符串会被当做字符串

#设置行高 列宽
ws = wb.create_sheet("设置行高与列宽")
ws.row_dimensions[2].height = 100#设置第二行的行高为100
ws.column_dimensions["E"].width = 500#设置第E列的列宽是50


wb.save("修改样式.xlsx")
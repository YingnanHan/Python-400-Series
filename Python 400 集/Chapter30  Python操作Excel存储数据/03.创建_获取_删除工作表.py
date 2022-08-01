import openpyxl
wb = openpyxl.load_workbook("excelTest.xlsx")
#根据工作簿创建工作表
wb.create_sheet(title="MySheet")
#获取所有工作表判断刚刚的创建是否成功
print(wb.sheetnames)
#根据工作表名称获取指定的工作表
my_sheet = wb.get_sheet_by_name("MySheet")
print(my_sheet)
#根据工作表名称删除工作表
wb.remove(my_sheet)
#获取所有工作表判断刚刚的删除是否成功
print(wb.sheetnames)
#获取工作表的方法②
#为了保证自洽重新创建MySheet工作表
wb.create_sheet(title="MySheet")
sheet = wb["MySheet"]
#输出工作表对象
print(sheet)
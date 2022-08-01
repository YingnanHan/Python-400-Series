import openpyxl

wb = openpyxl.load_workbook("work_book_excel.xlsx")

ws = wb.active

#①获取行
print("##获取第二行的数据##")
rows = ws[2]#得到第二行所有单元格的对象构成的元组
print(rows)
for r in rows:
    print(r.value,end="\t")
print()
#②获取列
print("##获取第二列的数据##")
cols = ws["B"]
print(cols)
for c in cols:
    print(c.value,end="\t")
print()
######对Excel-Sheet对象的切片保留的是闭区间######
#③获取部分行  例如 第二行以及第三行
print("##第二行以及第三行的数据##")
row_set = ws[2:3]
print(row_set)
for r in row_set:
    for c in r:
        print(c.value,end="\t")
    print()

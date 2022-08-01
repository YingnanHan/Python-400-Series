import openpyxl
wb = openpyxl.load_workbook("work_book_excel.xlsx")
ws =wb.active
#获取Excel中的sheet对象的行数以及列数的详细信息
print("共计{0}行，{1}列".format(ws.max_row,ws.max_column))

#①
#获取行的生成器对象
print(ws.iter_rows())
print("对全部行全部列的获取")
#使用iter_rows进行Excel sheet的遍历
for row in ws.iter_rows():
    for cell in row:
        print(cell.value,end="\t")
    print()
print()

#②
print()
print("对全部列全部行的获取")
#使用iter_rows进行Excel sheet的遍历
for col in ws.iter_cols():
    for cell in col:
        print(cell.value,end="\t")
    print()
print()


#③
#读取部分行部分列 使用iter_rows
print("对部分行部分列的获取：")
for row in ws.iter_rows(min_row=2,max_row=2,min_col=2,max_col=2):
    for c in row:#将范围缩小到第二行第二列的一个单元格里面
        print(c.value)

#和上面的区别是 行与列的先后顺序 （结果就相当于一个转置）

#④
#读取部分行部分列 使用iter_cols
print("对部分列部分行的获取：")
for row in ws.iter_cols(min_row=2,max_row=2,min_col=2,max_col=2):
    for c in row:#将范围缩小到第二行第二列的一个单元格里面
        print(c.value)


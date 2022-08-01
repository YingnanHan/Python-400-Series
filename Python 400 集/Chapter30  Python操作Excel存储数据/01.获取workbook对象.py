import openpyxl

#得到excel工作簿对象的两种方式

#1.创建excel文件 获取工作簿对象
excel1 = openpyxl.Workbook()

#2.加载已有的excel文件 获取工作簿对象
excel2 = openpyxl.load_workbook("excelTest.xlsx")

#3.对新创建的工作簿进行保存
excel1.save("work_book_excel.xlsx")

#4.获取所有工作表(sheet)的名称
print(excel1.sheetnames)#在操作excel文件的时候需要保证excel文件不被占用

#5.获取被激活的sheet名称
print(excel1.active)



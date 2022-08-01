import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

wb = openpyxl.Workbook()
ws = wb.active

#合并单元格
ws = wb.create_sheet("merged")
ws.merge_cells("A1:D3")
ws["A1"].value = "12 cells merged together"
ws.merge_cells("C5:D5")
ws["C5"].value = "2 cells merged together"

wb.save("styleexcelText.xlsx")
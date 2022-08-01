import openpyxl

wb = openpyxl.load_workbook("target.xlsx")
ws = wb.create_sheet("cell_sheet")

for row in range(10,21):
    for col in range(5,16):
        #指定行 指定列的写入操作
        ws.cell(row,col,row*col)

wb.save("target.xlsx")#当调用save函数之后会在原有的基础上新增加数据或者覆盖
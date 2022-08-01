import openpyxl

wb = openpyxl.load_workbook("work_book_excel.xlsx")

ws = wb.active

c = ws["C3"]
#①
#得到某一个单元格的行 列 及其对应的数值
print("row:{0},column:{1},value:{2}".format(c.row,c.column,c.value))
#②
#获取某一个单元格的坐标
print("coordinate:",c.coordinate)
#③
#获取某一个单元格
#WS.cell()获取单元格的value
print(ws.cell(1,2).value)
#④
print("\n")
print("循环遍历获取所有的元素:")
for r in ws.rows:
    for c in r:
        print(c.value,end="\t\t")
    print()
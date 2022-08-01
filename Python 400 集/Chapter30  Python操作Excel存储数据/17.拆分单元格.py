import openpyxl
from openpyxl.styles import Font
from openpyxl.styles import colors

wb = openpyxl.load_workbook("styleexcelText.xlsx")
ws = wb.active

#拆分为单元格 将merged sheet 复制一份并且保存
ws = wb.copy_worksheet(wb.get_sheet_by_name("merged"))
ws.title="unmerged"
ws.unmerge_cells("A1:D3")
ws.unmerge_cells("C5:D5")
wb.save("styleexcelText.xlsx")
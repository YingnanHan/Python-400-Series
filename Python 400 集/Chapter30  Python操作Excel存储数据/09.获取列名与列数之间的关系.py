import openpyxl
from openpyxl.utils import get_column_letter,column_index_from_string

#从数字映射字母
print("第2列对应的字母:",get_column_letter(2))
print("第100列对应的字母:",get_column_letter(100))
print("第1000列对应的字母:",get_column_letter(1000))

#从字母映射数字
print("字母D对应的列数:",column_index_from_string("D"))
print("字母DF对应的列数:",column_index_from_string("DF"))
print("字母DC对应的列数:",column_index_from_string("DC"))
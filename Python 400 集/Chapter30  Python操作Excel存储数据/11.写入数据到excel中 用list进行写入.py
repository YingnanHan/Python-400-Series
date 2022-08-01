import openpyxl

wb = openpyxl.Workbook()

ws = wb.create_sheet(title="使用List进行写入")
print(wb.sheetnames)

print("使用sheet进行写入")
rows = [
    ["now1","now2","now3","now4"],
    [12,34,56,78],
    [23,34,45,56],
    [34,45,56,67],
    [45,56,67,78]
]
#每一次写一行
for row in rows:
        ws.append(row)
wb.save("target.xlsx")
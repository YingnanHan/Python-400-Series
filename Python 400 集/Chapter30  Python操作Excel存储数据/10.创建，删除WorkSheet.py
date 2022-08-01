import openpyxl

#创建一个excel文件
wb = openpyxl.Workbook()
#激活一个excel文件中的sheet
ws = wb.active
#输出该sheet的名称
print(ws.title)
#修改工作簿的名称
ws.title = "My first work book"
#创建新的工作表
print("创建工作表")
wb.create_sheet(index=1,title="First Sheet")
wb.create_sheet(index=2,title="Second Sheet")
wb.create_sheet(index=3,title="Third Sheet")
print(wb.sheetnames)

#删除某一个工作表
print("删除工作表")#按照工作表名称进行删除
wb.remove(wb["Second Sheet"])
print(wb.sheetnames)